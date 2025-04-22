from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def format_excel(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column
        column_letter = get_column_letter(column)

    for column in sheet.columns:
        max_length = 0
        col = column[0].column  # Get column name (A, B, C...)
        column_letter = get_column_letter(col)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column_letter].width = max_length + 2

    new_path = file_path.replace(".xlsx", "_formatted.xlsx")
    wb.save(new_path)
    print(f"Formatted file saved as: {new_path}")


format_excel("data.xlsx")
